import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import random

# Data for realistic legal Q&A generation
supreme_court_data = {
    "cases": [
        "Kesavananda Bharati v. State of Kerala",
        "Indra Sawhney v. Union of India",
        "Golaknath v. State of Punjab",
        "Minerva Mills v. Union of India",
        "S.P. Gupta v. Union of India",
        "A.D.M. Jabalpur v. Shukla",
        "Maneka Gandhi v. Union of India",
        "Suresh Gupta v. Govt. of NCT Delhi",
        "Rajendra Prasad v. State of Uttar Pradesh",
        "M. Nagaraj v. Union of India"
    ],
    "topics": [
        "Constitutional validity",
        "Fundamental rights",
        "Due process",
        "Judicial review",
        "Electoral practices",
        "Right to property",
        "Right to life",
        "Freedom of expression",
        "Equal protection",
        "Executive powers"
    ],
    "questions": [
        "What was the primary issue in the case?",
        "Which constitutional article was violated?",
        "What was the court's judgment?",
        "What is the legal principle established?",
        "How does this ruling affect public policy?",
        "What was the dissenting opinion?",
        "What was the bench composition?",
        "Which articles of the Constitution were cited?",
        "What precedent did this case set?",
        "How long was the hearing?"
    ]
}

high_court_data = {
    "cases": [
        "State of Maharashtra v. Rajendra Singh",
        "Haryana State Electronics Development Corp. v. M/s Mittal Steel",
        "Delhi High Court v. Ministry of Finance",
        "Tamil Nadu v. Suresh Industries",
        "Karnataka High Court v. Bangalore Development Authority",
        "Gujarat High Court v. M/s Reliance Industries",
        "Punjab High Court v. State of Punjab",
        "West Bengal High Court v. Kolkata Municipal Corp.",
        "Bombay High Court v. State of Maharashtra",
        "Madras High Court v. Chennai Port Authority"
    ],
    "topics": [
        "Land acquisition",
        "Labor rights",
        "Environmental protection",
        "Revenue matters",
        "Consumer protection",
        "Administrative law",
        "Taxation",
        "Municipal governance",
        "Power distribution",
        "Contract enforcement"
    ],
    "questions": [
        "What was the dispute about?",
        "Which act was applicable?",
        "What compensation was awarded?",
        "What interim relief was granted?",
        "How was the evidence evaluated?",
        "What alternative remedies exist?",
        "How did the court interpret the statute?",
        "What procedural violations occurred?",
        "How was damages calculated?",
        "What was the legal standard applied?"
    ]
}

contract_data = {
    "types": [
        "Supply Agreement",
        "Service Agreement",
        "Purchase Agreement",
        "Licensing Agreement",
        "Partnership Agreement",
        "Employment Contract",
        "Lease Agreement",
        "NDA (Non-Disclosure Agreement)",
        "Sales Contract",
        "Joint Venture Agreement"
    ],
    "topics": [
        "Consideration",
        "Breach of contract",
        "Force majeure",
        "Liability",
        "Termination clause",
        "Payment terms",
        "Dispute resolution",
        "Confidentiality",
        "Indemnification",
        "Warranty and conditions"
    ],
    "questions": [
        "What are the key terms of the contract?",
        "How is the contract terminated?",
        "What happens in case of breach?",
        "How is payment structured?",
        "What are the liabilities?",
        "How are disputes resolved?",
        "What conditions must be met?",
        "How long is the contract valid?",
        "What are the confidentiality provisions?",
        "What warranties are provided?"
    ]
}

employment_data = {
    "contract_types": [
        "Permanent Employment",
        "Fixed Term Contract",
        "Probationary Employment",
        "Contractual Labor",
        "Senior Management",
        "Apprenticeship",
        "Internship",
        "Consultant Agreement",
        "Remote Work Agreement",
        "Executive Agreement"
    ],
    "topics": [
        "Salary and benefits",
        "Working hours",
        "Leave policy",
        "Termination conditions",
        "Non-compete clause",
        "Performance evaluation",
        "Health and safety",
        "Dispute resolution",
        "Retirement benefits",
        "Compensation for termination"
    ],
    "questions": [
        "What is the employment period?",
        "What benefits are included?",
        "How much notice is required for termination?",
        "What is the CTC?",
        "What are the job responsibilities?",
        "How is performance measured?",
        "What leave is entitled?",
        "Is there a non-compete clause?",
        "What is the retirement plan?",
        "How are disputes resolved?"
    ]
}

statute_data = {
    "acts": [
        "Indian Constitution",
        "Indian Penal Code",
        "Code of Civil Procedure",
        "Code of Criminal Procedure",
        "Patents Act, 1970",
        "Copyright Act, 1957",
        "Companies Act, 2013",
        "FEMA (Foreign Exchange Management Act)",
        "Data Protection Act",
        "Environmental Protection Act"
    ],
    "topics": [
        "Jurisdictional scope",
        "Penalty provisions",
        "Procedural requirements",
        "Exemptions",
        "Amendment provisions",
        "Enforcement mechanism",
        "Statute of limitations",
        "Burden of proof",
        "Definitions",
        "Application and scope"
    ],
    "questions": [
        "What is the scope of this act?",
        "Who is authorized to enforce this?",
        "What penalties are provided?",
        "How is this act interpreted?",
        "What are the exemptions?",
        "When did this act come into force?",
        "How can this act be amended?",
        "What procedures must be followed?",
        "What is the limitation period?",
        "How is compliance verified?"
    ]
}

def generate_qa_pair(doc_type, index):
    """Generate a realistic Q&A pair based on document type"""
    
    if doc_type == "Supreme Court":
        case = random.choice(supreme_court_data["cases"])
        topic = random.choice(supreme_court_data["topics"])
        question_template = random.choice(supreme_court_data["questions"])
        
        answers = {
            "What was the primary issue in the case?": f"The case involved {topic.lower()} under the Constitution of India.",
            "Which constitutional article was violated?": f"Articles relating to {topic.lower()} were at stake.",
            "What was the court's judgment?": f"The court upheld the principles of {topic.lower()} in light of constitutional provisions.",
            "What is the legal principle established?": f"This case established that {topic.lower()} must be protected uniformly.",
            "How does this ruling affect public policy?": f"The ruling strengthens existing policies on {topic.lower()}.",
            "What was the dissenting opinion?": f"The dissent argued that {topic.lower()} should be interpreted differently.",
            "What was the bench composition?": "A constitution bench of 5 judges heard the matter.",
            "Which articles of the Constitution were cited?": f"Articles 14, 19, 21 and relevant provisions on {topic.lower()} were cited.",
            "What precedent did this case set?": f"This case set precedent for all future matters concerning {topic.lower()}.",
            "How long was the hearing?": "The matter was heard over multiple days with extensive arguments."
        }
        answer = answers.get(question_template, f"The Supreme Court addressed issues of {topic.lower()}.")
        
        return case, question_template, answer
    
    elif doc_type == "High Court":
        case = random.choice(high_court_data["cases"])
        topic = random.choice(high_court_data["topics"])
        question_template = random.choice(high_court_data["questions"])
        
        answers = {
            "What was the dispute about?": f"The dispute concerned {topic.lower()} between the parties.",
            "Which act was applicable?": f"The {topic} Act and relevant state laws were applicable.",
            "What compensation was awarded?": f"Compensation was awarded based on valuation of {topic.lower()}.",
            "What interim relief was granted?": f"The court granted interim relief to protect {topic.lower()} interests.",
            "How was the evidence evaluated?": f"Documentary and oral evidence regarding {topic.lower()} was examined.",
            "What alternative remedies exist?": "The party can approach the Supreme Court or seek review.",
            "How did the court interpret the statute?": f"The statute was interpreted to encompass {topic.lower()} matters.",
            "What procedural violations occurred?": "Procedural compliance was verified throughout the proceedings.",
            "How was damages calculated?": "Damages were calculated based on actual loss and projected impact.",
            "What was the legal standard applied?": f"The preponderance of probabilities standard was applied for {topic.lower()}."
        }
        answer = answers.get(question_template, f"The High Court decided on {topic.lower()} matters.")
        
        return case, question_template, answer
    
    elif doc_type == "Commercial Contract":
        contract_type = random.choice(contract_data["types"])
        topic = random.choice(contract_data["topics"])
        question_template = random.choice(contract_data["questions"])
        
        answers = {
            "What are the key terms of the contract?": f"Payment: Fixed amount, Delivery: Within 30 days, {topic}: As per contract schedule.",
            "How is the contract terminated?": f"The contract can be terminated with 30 days notice upon {topic} conditions.",
            "What happens in case of breach?": f"A breach of {topic} entitles the non-breaching party to damages or specific performance.",
            "How is payment structured?": f"Payment is made in milestones as per {topic} provisions: 40% upfront, 60% on delivery.",
            "What are the liabilities?": f"Each party's liability is capped at the contract value except for {topic} matters.",
            "How are disputes resolved?": f"Disputes are resolved through arbitration under {topic} rules.",
            "What conditions must be met?": f"All conditions regarding {topic} must be satisfied for liability release.",
            "How long is the contract valid?": "The contract is valid for 3 years from the date of execution.",
            "What are the confidentiality provisions?": f"Parties must maintain confidentiality on {topic} for 5 years post-termination.",
            "What warranties are provided?": f"The supplier warrants {topic} compliance for 12 months from delivery."
        }
        answer = answers.get(question_template, f"The contract specifies terms for {topic.lower()}.")
        
        return f"{contract_type} - Case #{index}", question_template, answer
    
    elif doc_type == "Employment Contract":
        employment_type = random.choice(employment_data["contract_types"])
        topic = random.choice(employment_data["topics"])
        question_template = random.choice(employment_data["questions"])
        
        answers = {
            "What is the employment period?": f"The employment is for a fixed period of {random.randint(1, 5)} years.",
            "What benefits are included?": f"Medical insurance, annual bonus, and {topic} benefits as per company policy.",
            "How much notice is required for termination?": f"Either party must provide 3 months notice, except in cases of {topic} violations.",
            "What is the CTC?": f"The CTC is determined based on experience and {topic} considerations.",
            "What are the job responsibilities?": f"The employee is responsible for {topic.lower()} related duties and functions.",
            "How is performance measured?": f"Performance is measured through quarterly reviews with {topic} metrics.",
            "What leave is entitled?": "Annual leave: 20 days, Sick leave: 10 days, Special leave: As per company policy.",
            "Is there a non-compete clause?": f"Yes, effective for 2 years post-employment regarding {topic.lower()}.",
            "What is the retirement plan?": f"Retirement benefits include gratuity and pension as per {topic} norms.",
            "How are disputes resolved?": f"Disputes related to {topic.lower()} are resolved through arbitration."
        }
        answer = answers.get(question_template, f"The employment contract specifies {topic.lower()} terms.")
        
        return f"{employment_type} - Emp #{index}", question_template, answer
    
    else:  # Legal Statute
        act = random.choice(statute_data["acts"])
        topic = random.choice(statute_data["topics"])
        question_template = random.choice(statute_data["questions"])
        
        answers = {
            "What is the scope of this act?": f"This act applies to {topic.lower()} matters across all states and territories.",
            "Who is authorized to enforce this?": f"The designated authority and government bodies are responsible for {topic.lower()} enforcement.",
            "What penalties are provided?": f"Penalties for {topic.lower()} violations range from fines to imprisonment.",
            "How is this act interpreted?": f"The act is interpreted based on its literal meaning and legislative intent regarding {topic.lower()}.",
            "What are the exemptions?": f"Exemptions exist for specific categories as defined under {topic.lower()} provisions.",
            "When did this act come into force?": f"This act regarding {topic.lower()} came into force from the specified date.",
            "How can this act be amended?": f"Amendments to {topic.lower()} provisions require legislative approval.",
            "What procedures must be followed?": f"Detailed procedures for {topic.lower()} compliance are outlined in the rules.",
            "What is the limitation period?": f"The limitation period for {topic.lower()} actions is typically 3 years.",
            "How is compliance verified?": f"Compliance with {topic.lower()} standards is verified through regular audits."
        }
        answer = answers.get(question_template, f"The {act} specifies {topic.lower()} requirements.")
        
        return act, question_template, answer

def create_test_dataset():
    """Create the complete test dataset"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Dataset"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 15
    
    # Create headers
    headers = ["Case/Document Name", "Document Type", "Question", "Answer", "Difficulty"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Distribution based on test dataset
    distribution = {
        "Supreme Court": 350,
        "High Court": 210,
        "Commercial Contract": 175,
        "Employment Contract": 140,
        "Legal Statute": 125
    }
    
    row = 2
    difficulties = ["Easy", "Medium", "Hard"]
    
    # Generate Q&A pairs
    for doc_type, count in distribution.items():
        for idx in range(1, count + 1):
            case_name, question, answer = generate_qa_pair(doc_type, idx)
            difficulty = random.choice(difficulties)
            
            ws.cell(row=row, column=1, value=case_name)
            ws.cell(row=row, column=2, value=doc_type)
            ws.cell(row=row, column=3, value=question)
            ws.cell(row=row, column=4, value=answer)
            ws.cell(row=row, column=5, value=difficulty)
            
            # Apply text wrapping for better readability
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            
            row += 1
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 15
    
    # Summary header
    summary_ws['A1'] = "Document Type"
    summary_ws['B1'] = "Question Count"
    for cell in ['A1', 'B1']:
        summary_ws[cell].fill = header_fill
        summary_ws[cell].font = header_font
    
    summary_row = 2
    total_questions = 0
    for doc_type, count in distribution.items():
        summary_ws.cell(row=summary_row, column=1, value=doc_type)
        summary_ws.cell(row=summary_row, column=2, value=count)
        total_questions += count
        summary_row += 1
    
    summary_ws.cell(row=summary_row, column=1, value="TOTAL")
    summary_ws.cell(row=summary_row, column=2, value=total_questions)
    summary_ws.cell(row=summary_row, column=1).font = Font(bold=True)
    summary_ws.cell(row=summary_row, column=2).font = Font(bold=True)
    summary_ws.cell(row=summary_row, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    summary_ws.cell(row=summary_row, column=2).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Save file
    output_file = "e:\\kanunai\\legal_qa_test_dataset_1000.xlsx"
    wb.save(output_file)
    print(f"✓ Test dataset created successfully!")
    print(f"✓ Total questions: {total_questions}")
    print(f"✓ File saved: {output_file}")
    print(f"\nDistribution:")
    for doc_type, count in distribution.items():
        print(f"  - {doc_type}: {count} questions")

if __name__ == "__main__":
    create_test_dataset()
